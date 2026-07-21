"""模範トークスクリプト（Googleスプレッドシート）を KNOTE に取り込む。

1人ロープレの「🎯模範トーク視点」の採点基準になる社内標準スクリプト。
複数ブック・複数タブをまとめて1本のテキストに統合して保存する。

- 読み取りは DWD サービスアカウント（drive.readonly）で、ブックを xlsx で書き出して解析。
  ※Sheets スコープが委任されていないため、Drive のエクスポート経由で取得する。
- 保存先は storage の script セクション（KNOWLEDGE_SHEET_ID 設定時は共有シートの
  TalkScript タブ）。スクリプトが増えたら再実行すれば丸ごと更新される。

使い方:
    python3 scripts/import_talk_scripts.py [--dry-run]
"""
from __future__ import annotations

import argparse
import io
import re
import sys
import urllib.request

from google.auth.transport.requests import Request
from google.oauth2 import service_account

from config import settings
from services import storage

# 取り込む模範トークスクリプトのブック（タイトルはログ表示用）
SCRIPT_BOOKS = [
    ("1bUrsQxKIg7DZc-DANvHRqR4ygeEQEOvckZTpeGlmyEs", "商材別トークスクリプト"),
    ("1A92RANfvc9zQz18b9aG4fnSlurKAqBVr0iWAzGSAX1k", "説明トーク（部品別）"),
]
SUBJECT = "planner@life-time-support.com"
HEADER = "===== 模範トークスクリプト（社内標準） ====="


def _token() -> str:
    if not settings.GOOGLE_SERVICE_ACCOUNT_FILE:
        sys.exit("GOOGLE_SERVICE_ACCOUNT_FILE が未設定です。")
    creds = service_account.Credentials.from_service_account_file(
        settings.GOOGLE_SERVICE_ACCOUNT_FILE,
        scopes=["https://www.googleapis.com/auth/drive.readonly"],
        subject=SUBJECT,
    )
    creds.refresh(Request())
    return creds.token


def _is_heading(cell) -> bool:
    """セルに色が塗ってあれば『項目の見出し』とみなす（黄色・赤/ピンク等）。"""
    fill = getattr(cell, "fill", None)
    if fill is None or getattr(fill, "patternType", None) != "solid":
        return False
    rgb = getattr(getattr(fill, "fgColor", None), "rgb", None)
    return bool(rgb) and rgb not in ("00000000", "FFFFFFFF")


def _book_text(sheet_id: str, token: str) -> tuple[list[tuple[str, str]], int]:
    """ブックを xlsx で取得し、[(タブ名, 本文)] と合計文字数を返す。

    - セル1つ＝1段落として改行し、読みやすく整形する。
    - 色付きセル（黄色・赤等）は項目の見出しなので【】で強調し、前に空行を入れる。
    """
    import openpyxl

    req = urllib.request.Request(
        f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx",
        headers={"Authorization": f"Bearer {token}"},
    )
    wb = openpyxl.load_workbook(io.BytesIO(urllib.request.urlopen(req, timeout=120).read()),
                                data_only=True)
    out: list[tuple[str, str]] = []
    total = 0
    for name in wb.sheetnames:
        ws = wb[name]
        # 先に見出し（色付きセル）の文言を集める。別列にある同じ文字列は
        # 目次（インデックス列）なので本文からは除く。
        headings = {
            str(c.value).strip()
            for row in ws.iter_rows() for c in row
            if c.value and str(c.value).strip() and _is_heading(c)
        }
        lines: list[str] = []
        for row in ws.iter_rows():
            for cell in row:
                val = cell.value
                if val is None or not str(val).strip():
                    continue
                text = str(val).strip()
                if _is_heading(cell):
                    lines.append("")           # 見出しの前に空行
                    lines.append(f"【{text}】")
                elif text in headings:
                    continue                   # 目次列の重複は捨てる
                else:
                    lines.append(text)
        body = "\n".join(lines).strip()
        body = re.sub(r"\n{3,}", "\n\n", body)   # 空行の詰めすぎを整える
        if body:
            out.append((name, body))
            total += len(body)
    return out, total


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    token = _token()
    items: list[dict] = []
    grand = 0
    for sheet_id, title in SCRIPT_BOOKS:
        tabs, total = _book_text(sheet_id, token)
        grand += total
        print(f"■ {title}: {len(tabs)}タブ / {total:,}字")
        for name, body in tabs:
            print(f"    - {name} … {len(body):,}字")
            items.append({"book": title, "tab": name, "body": body})

    print(f"\n単元数: {len(items)} / 合計 {grand:,} 文字（≈ {grand//3:,} tokens 目安）")
    if args.dry_run:
        print("--dry-run のため保存しません。")
        return
    # 単元（元スプシのタブ）ごとに1行で保存 → 中身をタブ単位で確認・修正できる
    storage.set_talk_script_items(items)
    where = "共有シートの TalkScripts タブ" if storage._use_sheets() else "ローカル"
    print(f"💾 保存しました → {where}"
          f"（読み戻し {len(storage.get_talk_script_items())} 単元 / "
          f"連結 {len(storage.get_talk_script()):,} 字）")


if __name__ == "__main__":
    main()
