"""模範トークスクリプトの各タブから、1人ロープレの「商材・単元別シナリオ」を作る。

- スクリプト1タブ ＝ ロープレ1単元（例：ガラス説明／セラミック説明／エコカラットの価格…）。
- 各ターンの「カンペ」は **実際のスクリプト本文** をそのまま使う（AI不使用＝無料）。
- お客様のセリフは、タブ名と定型の掘り下げ質問から機械的に組み立てる。
- 手書きの総合シナリオ（初回商談／反論対応／クロージング）は先頭に残す。

使い方:
    python3 scripts/build_roleplay_scenarios.py [--dry-run] [--max-turns 4]
"""
from __future__ import annotations

import argparse
import io
import re
import sys
import urllib.request

from google.auth.transport.requests import Request
from google.oauth2 import service_account

from config import settings
from services import storage

SCRIPT_BOOKS = [
    "1bUrsQxKIg7DZc-DANvHRqR4ygeEQEOvckZTpeGlmyEs",
    "1A92RANfvc9zQz18b9aG4fnSlurKAqBVr0iWAzGSAX1k",
]
SUBJECT = "planner@life-time-support.com"

# 対象カテゴリは4つに限定（重点3商材＋導入）。他商材は将来実装のため今は生成しない。
# 判定はこの順で行う（コーティング等を先に判定し、汎用の「導入」は最後に拾う）。
CATEGORIES = [
    ("コーティング", ["コーティング", "ガラス説明", "ＵＶ説明", "UV説明", "シリコン説明",
                     "セラミック説明", "整い", "水回りコーティング", "水廻り", "フロアコーティング"]),
    ("エコカラット", ["エコカラット", "エコＬＤ", "エコLD"]),
    ("ダウンライト", ["ダウンライト", "人感センサー", "照明"]),
    ("導入", ["導入", "会社説明", "内覧会", "締め", "スケジュール", "引っ越し", "引越",
             "検討内容", "商談の流れ"]),
]


def _category_of(tab: str) -> str | None:
    for name, keys in CATEGORIES:
        if any(k in tab for k in keys):
            return name
    return None  # 4カテゴリ外（エアコン・家具・住設等）は対象外


# ロープレ単元から除外するタブ（重複・統合版のため個別単元にしない）。先頭「未」除去後の名前で判定。
UNIT_EXCLUDE = {"整い版", "コーティング"}

# カテゴリ内の単元の並び順（キーワードで先頭一致を判定。未該当は末尾）。
CATEGORY_UNIT_ORDER = {
    "導入": ["ラポール", "商談の導入", "会社説明", "スケジュール", "内覧会", "締め", "引っ越し"],
    "コーティング": [
        "フロアコーティングを行った", "種類を話す前の導入", "ガラス", "ＵＶ", "UV",
        "シリコン", "セラミック", "水回り", "水廻り", "特典", "決まった後",
    ],
}

# ラポール単元の評価観点：手順ではなく“人としての魅力と信頼”が伝わったかを見る。
RAPPORT_FOCUS = "\n".join([
    "この単元は『ラポールの形成』。評価の中心は、お客様が",
    "『この人、話していて楽しい・面白い』『この人は信頼できる』と感じられたか。",
    "- 親しみ・温かさ：笑顔が伝わる言葉選び、相手への素直な関心、会話の楽しさ（一方通行でない）",
    "- 信頼感：誠実さ・プロとしての安心感、押し売り感のなさ、説明の明確さ",
    "- 相手を主役にする：自分が話しすぎず、お客様に気持ちよく話させたか",
    "- 自然な流れ：挨拶→雑談→商談へ、ぶつ切りでなく1つのストーリーとして繋がっているか",
    "※単に手順を踏んだかではなく、『人としての魅力と信頼』が伝わったかで採点する。",
    "いきなり商材説明に入る・自己紹介や雑談が端的すぎる場合は、内容が正しくても高評価にしない。",
])

# 手作りの段階的ショートロープレ：導入【ラポールの形成】（挨拶→雑談→自然に商談へ）。
RAPPORT_TURNS = [
    {"customer": "（着席して）本日はよろしくお願いします。",
     "hint": "【STEP1 第一印象・自己紹介】\n明るく笑顔で名乗り、時間をいただいた感謝を伝える。\n"
             "モデル：『初めまして！ライフタイムサポートの◯◯と申します。本日はお忙しい中、"
             "お時間をいただきありがとうございます。』\n→ 早口・事務的にならない。まず“感じの良い人”という第一印象を作る。"},
    {"customer": "よろしくお願いします。（少し緊張している）",
     "hint": "【STEP2 雑談・アイスブレイク】\nいきなり商材に入らない。物件・アクセス・引越しの話題で緊張をほぐす。\n"
             "切り口例：『新しいお住まい、素敵ですね！ご入居が楽しみですね。』"
             "『こちらまで迷わず来られましたか？』\n→ お客様が一言でも“自分の話”をしたらラポールの入口。"},
    {"customer": "はい。オプションとか、正直よく分かってなくて…。",
     "hint": "【STEP3 今日のゴール共有・安心づけ】\n“決めさせる場ではない”と伝えて警戒を解く。\n"
             "モデル：『大丈夫です、皆さん最初はそうです。今日は決める場ではなく、どんなことができるかを"
             "知っていただく場だと思ってください。実際に決めるのは内覧会の後で、それまでゆっくり悩めます。』"},
    {"customer": "そうなんですね、それなら安心です。",
     "hint": "【STEP4 会社説明を“物語として・端的すぎず”】\n押し付けず、弊社の強みを1〜2点だけ自然に。\n"
             "切り口例：『弊社はもともとマンションのオプション工事を請けていた会社なので、オプション会で"
             "紹介されるものは基本何でもできます。中間マージンが省ける分お得なんです。』"
             "『内覧会に無料同行してお部屋を一緒にチェックするサービスも好評です。』\n"
             "→ “だから何でも気軽に相談してくださいね”に着地させる。"},
    {"customer": "はい、色々聞いてみたいです。",
     "hint": "【STEP5 自然に商談（ヒアリング）へ移行】\n雑談から本題へ橋渡し。質問で主導権をお客様に渡す。\n"
             "切り口例：『では、今日は気になっていることを何でも聞かせてください。エコカラットやコーティングなど、"
             "検討されているものはありますか？“ここ不便だな”と感じることでも大丈夫です。』\n"
             "→ ここで出た関心・困りごとが、この後の商材提案の土台になる。"},
]


def _unit_rank(cat: str, title: str) -> tuple:
    keys = CATEGORY_UNIT_ORDER.get(cat)
    if keys:
        for i, k in enumerate(keys):
            if k in title:
                return (i, "")
        return (len(keys), title)
    return (0, title)

FOLLOW_UPS = [
    "なるほど。もう少し詳しく教えてもらえますか？",
    "それだと、どんなメリットがあるんでしょうか。",
    "他の選択肢と比べると、どう違いますか？",
    "費用面や期間はどうなりますか？",
]


def _opening_line(tab: str) -> str:
    t = tab.strip()
    if t.endswith("？") or t.endswith("?"):
        return t
    t = re.sub(r"^(未|必ず話す)\s*", "", t)
    return f"{t}について、まだよく分かっていなくて。教えてもらえますか？"


# 「商談の導入」系は、信頼関係づくりを最重視して採点させる
TRUST_FOCUS = (
    "この単元の目的は『お客様と信頼関係を築き、なんでも相談してもらえる雰囲気をつくる』こと。"
    "次を最重視して採点する：\n"
    "- 警戒をほどく共感・受け止め（否定しない・今すぐ決めさせない）\n"
    "- 今日のゴール共有で安心させたか（『今日は決める場ではない』と伝えられたか）\n"
    "- 相談しやすい問いかけ（オープンな質問で、お客様が話す余白をつくれたか）\n"
    "- 売り込み感の排除（提案より先に“理解しようとする姿勢”が出ていたか）\n"
    "- 結果として、お客様が本音や不安を漏らし始める雰囲気になったか\n"
    "※ここが弱ければ、商品説明がどれだけ流暢でも高評価にしないこと。"
)
TRUST_KEYS = ["導入", "会社説明", "スケジュール"]

# お客様の属性別フォーカス（段階的にスキルアップさせる）
UNDECIDED_FOCUS = "\n".join([
    "【お客様属性：未検討】まだ必要性を感じていないお客様。ここが最難関で、最重要。",
    "- いきなり商品説明に入っていないか（説明先行は減点）",
    "- 暮らし方・家族・不満・将来の困りごとを聞き出し、本人に『言わせた』か",
    "- 引き出したニーズと商材を結び付けられたか（機能の羅列で終わっていないか）",
    "- 結果として『それなら必要かも』と本人が納得する流れを作れたか",
])
DECIDED_FOCUS = "\n".join([
    "【お客様属性：検討済み】すでに検討・見積取得済みのお客様。",
    "- 既存の検討内容を否定せず受け止めたうえで、まだ気づいていない観点を足せたか",
    "- 価格比較だけの土俵から、中身（耐久・保証・範囲）の比較へ移せたか",
    "- 見積から外されないよう、必要性を本人の言葉で再確認できたか",
])


def _customer_type(tab: str) -> tuple[str, int]:
    """タブ名から お客様属性 と 難易度レベル を決める。『未〜』は未検討版。"""
    if tab.strip().startswith("未"):
        return "未検討", 2
    return "検討済み", 1


def _sections(body: str, max_turns: int) -> list[tuple[str, str]]:
    """本文を【見出し】単位の章に分ける。戻り値 [(見出し, 本文)]。

    スプレッドシートの色付きセル＝項目見出しを【】で取り込んでいるので、
    その章立てをそのままロープレのターンに使う（読みやすく・練習しやすい）。
    """
    parts = re.split(r"\n?【([^】]+)】\n?", body)
    secs: list[tuple[str, str]] = []
    if len(parts) > 1:
        head = parts[0].strip()
        if head:
            secs.append(("導入部", head))
        for i in range(1, len(parts) - 1, 2):
            title, chunk = parts[i].strip(), parts[i + 1].strip()
            if chunk:
                secs.append((title, chunk))
    else:
        paras = [p.strip() for p in re.split(r"\n\s*\n", body) if p.strip()]
        if not paras:
            return []
        size = max(1, (len(paras) + max_turns - 1) // max_turns)
        secs = [("", "\n".join(paras[i:i + size])) for i in range(0, len(paras), size)]
    return secs[:max_turns]


def _token() -> str:
    if not settings.GOOGLE_SERVICE_ACCOUNT_FILE:
        sys.exit("GOOGLE_SERVICE_ACCOUNT_FILE が未設定です。")
    creds = service_account.Credentials.from_service_account_file(
        settings.GOOGLE_SERVICE_ACCOUNT_FILE,
        scopes=["https://www.googleapis.com/auth/drive.readonly"],
        subject=SUBJECT,
    )
    creds.refresh(Request())
    return creds.token


def _tabs(sheet_id: str, token: str) -> list[tuple[str, str]]:
    import openpyxl

    req = urllib.request.Request(
        f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx",
        headers={"Authorization": f"Bearer {token}"},
    )
    wb = openpyxl.load_workbook(io.BytesIO(urllib.request.urlopen(req, timeout=120).read()),
                                data_only=True)
    out = []
    for name in wb.sheetnames:
        cells = []
        for row in wb[name].iter_rows(values_only=True):
            for c in row:
                if c and str(c).strip():
                    cells.append(str(c).strip())
        body = "\n".join(cells).strip()
        if body:
            out.append((name, body))
    return out


def _downlight_units(items: list[dict]) -> list[tuple[str, str]]:
    """住設タブから、ダウンライト/照明/人感センサーの範囲を切り出して単元化する。

    ダウンライト専用タブが無いため、住設の「人感センサー」〜「壁掛け」手前を抜き出す。
    住設タブはほぼ全セルが色付きで見出し判定が過剰なため、【】は一旦解除して整える。
    """
    body = next((i.get("body", "") for i in items if i.get("tab") == "住設"), "")
    if not body:
        return []
    m = re.search(r"【?人感センサー】?", body)
    if not m:
        return []
    seg = body[m.start():]
    end = re.search(r"【?壁掛け】?", seg)
    if end:
        seg = seg[: end.start()]
    seg = seg.replace("【", "").replace("】", "").strip()
    seg = re.sub(r"\n{2,}", "\n", seg)
    return [("ダウンライト・人感センサー照明", "【ダウンライト・人感センサー・照明】\n" + seg)]


def _build_turns(tab: str, parts: list, ctype: str) -> list[dict]:
    turns = []
    for i, (title, chunk) in enumerate(parts):
        if i == 0:
            customer = _opening_line(tab)  # 第一声は render 時にペルソナで属性別に差し替わる
        elif title and title != "導入部":
            customer = f"（{title}について）もう少し教えてもらえますか？"
        else:
            customer = FOLLOW_UPS[(i - 1) % len(FOLLOW_UPS)]
        turns.append({"customer": customer, "hint": chunk, "section": title})
    return turns


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--max-turns", type=int, default=5)
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    items = storage.get_talk_script_items()
    source = [(i.get("tab", ""), i.get("body", "")) for i in items]
    if not source:
        sys.exit("模範トークスクリプトが未登録です。先に import_talk_scripts.py を実行してください。")

    # 4カテゴリに該当するタブだけを単元にする（＋ダウンライトは住設から抽出）
    units: list[tuple[str, str, str]] = []  # (category, title, body)
    seen_titles = set()
    for tab, body in source:
        cat = storage.script_category(tab)  # 編集画面と同じ分類（唯一の正）
        if cat == "その他" or not body.strip():
            continue
        if re.sub(r"^未", "", tab).strip() in UNIT_EXCLUDE:  # 整い版・コーティング(統合版)は除外
            continue
        key = (cat, tab)
        if key in seen_titles:
            continue
        seen_titles.add(key)
        units.append((cat, tab, body))
    for title, body in _downlight_units(items):
        units.append(("ダウンライト", title, body))

    # 各単元について、お客様タイプ2種（検討済み／未検討）を生成する
    scenarios: list[dict] = []
    used_ids = set()
    for cat, tab, body in units:
        parts = _sections(body, args.max_turns)
        if not parts:
            continue
        base = re.sub(r"[^0-9A-Za-zぁ-んァ-ン一-龥]+", "_", tab).strip("_")[:36] or "unit"
        # 表示タイトルは先頭の「未」（未商談用の意味）を外す（属性は別プルダウンで選ぶため）
        title = re.sub(r"^未", "", tab).strip()
        for ctype in ("検討済み", "未検討"):
            sid = f"{base}_{'d' if ctype == '検討済み' else 'u'}"
            while sid in used_ids:
                sid += "_x"
            used_ids.add(sid)
            focus = [UNDECIDED_FOCUS if ctype == "未検討" else DECIDED_FOCUS]
            if cat == "導入":
                focus.append(TRUST_FOCUS)
            scenarios.append({
                "id": sid, "group": cat, "title": title,
                "customer_type": ctype, "level": 2 if ctype == "未検討" else 1,
                "focus": "\n\n".join(focus),
                "turns": _build_turns(tab, parts, ctype),
            })

    # 導入【ラポールの形成】を追加（手作りの段階ロープレ・第一声は差し替えない）
    for ctype in ("検討済み", "未検討"):
        scenarios.append({
            "id": f"rapport_{'d' if ctype == '検討済み' else 'u'}",
            "group": "導入", "title": "ラポールの形成",
            "customer_type": ctype, "level": 1, "keep_opening": True,
            "focus": RAPPORT_FOCUS,
            "turns": [dict(t) for t in RAPPORT_TURNS],
        })

    # カテゴリの表示順を固定し、カテゴリ内は指定の単元順に並べる
    order = {"導入": 0, "コーティング": 1, "エコカラット": 2, "ダウンライト": 3}
    scenarios.sort(key=lambda s: (order.get(s["group"], 9),
                                  _unit_rank(s["group"], s["title"]), s["customer_type"]))

    from collections import Counter
    by_cat = Counter(s["group"] for s in scenarios)
    print(f"生成シナリオ: {len(scenarios)} 件（4カテゴリ × 2タイプ）")
    for c in ["導入", "コーティング", "エコカラット", "ダウンライト"]:
        titles = sorted({s["title"] for s in scenarios if s["group"] == c})
        print(f"  ■ {c}: {by_cat.get(c,0)}件 / 単元 {len(titles)}: {'、'.join(titles)}")

    if args.dry_run:
        print("--dry-run のため保存しません。")
        return

    # リセット→再構築（既存シナリオを置き換える）
    storage.set_scenarios(scenarios)
    print(f"💾 保存しました（読み戻し {len(storage.get_scenarios())} 件）")


if __name__ == "__main__":
    main()
